import io
from datetime import date

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from app.database import get_db
from app.schemas import (
    HoldingCreate,
    HoldingUpdate,
    HoldingResponse,
)

router = APIRouter(prefix="/users/{user_id}/portfolio", tags=["portfolio"])


@router.get("/template")
def download_template(user_id: int):
    """Download a blank Excel template for bulk portfolio import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Holdings"

    headers = ["Ticker", "Shares", "Avg Cost Per Share ($)"]
    header_fill = PatternFill("solid", fgColor="16A34A")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Example rows so the user understands the format
    ws.append(["AAPL", 10, 175.50])
    ws.append(["TSLA", 5, 220.00])

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=portfolio_template.xlsx"},
    )


@router.post("/bulk", status_code=201)
def bulk_add_holdings(user_id: int, file: UploadFile = File(...)):
    """Upload a filled-in Excel template to add multiple holdings at once."""
    import openpyxl

    with get_db() as conn:
        _require_user(conn, user_id)

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the uploaded file. Make sure it is a valid .xlsx file.")

    added, skipped, errors = 0, 0, []
    today = date.today().isoformat()

    with get_db() as conn:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue  # skip blank rows
            if len(row) < 3:
                errors.append(f"Row {row_idx}: not enough columns.")
                skipped += 1
                continue

            ticker_raw, shares_raw, cost_raw = row[0], row[1], row[2]

            if not ticker_raw:
                skipped += 1
                continue

            ticker = str(ticker_raw).strip().upper()
            try:
                shares = float(shares_raw)
                cost = float(cost_raw)
            except (TypeError, ValueError):
                errors.append(f"Row {row_idx} ({ticker}): Shares and cost must be numbers.")
                skipped += 1
                continue

            if shares <= 0 or cost < 0:
                errors.append(f"Row {row_idx} ({ticker}): Shares must be > 0 and cost >= 0.")
                skipped += 1
                continue

            conn.execute(
                """INSERT INTO holdings (user_id, ticker, shares, avg_cost_basis, acquired_date)
                   VALUES (?, ?, ?, ?, ?)""",
                (user_id, ticker, shares, cost, today),
            )
            added += 1

    return {"added": added, "skipped": skipped, "errors": errors}


@router.post("", response_model=HoldingResponse, status_code=201)
def add_holding(user_id: int, holding: HoldingCreate):
    with get_db() as conn:
        _require_user(conn, user_id)
        cursor = conn.execute(
            """INSERT INTO holdings
               (user_id, ticker, shares, avg_cost_basis, acquired_date)
               VALUES (?, ?, ?, ?, ?)""",
            (
                user_id,
                holding.ticker.upper(),
                holding.shares,
                holding.avg_cost_basis,
                holding.acquired_date,
            ),
        )
        row = conn.execute(
            "SELECT * FROM holdings WHERE id = ?", (cursor.lastrowid,)
        ).fetchone()
    return _holding_from_row(row)


@router.put("/{holding_id}", response_model=HoldingResponse)
def update_holding(user_id: int, holding_id: int, update: HoldingUpdate):
    with get_db() as conn:
        _require_user(conn, user_id)
        row = conn.execute(
            "SELECT * FROM holdings WHERE id = ? AND user_id = ?",
            (holding_id, user_id),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Holding not found")

        fields = {}
        if update.ticker is not None:
            fields["ticker"] = update.ticker.upper()
        if update.shares is not None:
            fields["shares"] = update.shares
        if update.avg_cost_basis is not None:
            fields["avg_cost_basis"] = update.avg_cost_basis
        if update.acquired_date is not None:
            fields["acquired_date"] = update.acquired_date

        if not fields:
            raise HTTPException(status_code=400, detail="No fields to update")

        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [holding_id, user_id]
        conn.execute(
            f"UPDATE holdings SET {set_clause} WHERE id = ? AND user_id = ?",
            values,
        )
        updated = conn.execute(
            "SELECT * FROM holdings WHERE id = ?", (holding_id,)
        ).fetchone()
    return _holding_from_row(updated)


@router.delete("/{holding_id}", status_code=204)
def delete_holding(user_id: int, holding_id: int):
    with get_db() as conn:
        _require_user(conn, user_id)
        row = conn.execute(
            "SELECT id FROM holdings WHERE id = ? AND user_id = ?",
            (holding_id, user_id),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Holding not found")
        conn.execute("DELETE FROM holdings WHERE id = ?", (holding_id,))
    return None


def _require_user(conn, user_id: int) -> None:
    row = conn.execute(
        "SELECT id FROM users WHERE id = ?", (user_id,)
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="User not found")


def _holding_from_row(row: dict) -> HoldingResponse:
    return HoldingResponse(
        id=row["id"],
        user_id=row["user_id"],
        ticker=row["ticker"],
        shares=row["shares"],
        avg_cost_basis=row["avg_cost_basis"],
        acquired_date=row["acquired_date"],
        created_at=row["created_at"],
    )
